#!/usr/bin/python3

from pprint import pprint
from optparse import OptionParser
import pandas as pd
import json
import re
import os
from openpyxl import load_workbook
from datetime import datetime

class Survey:

    def __init__(self, date, df = None, name = "base"):
        """

        :param date: Date of the current report
        """
        self.date = date
        self.df   = df
        self.name = name

    def read_data_from_csv(self, path, filename):

        df = pd.read_csv(os.path.join(path,filename))

        # Remove SurveyMonkey columns not required
        df = df.drop(columns=['respondent_id',
                            'collector_id',
                            'date_created',
                            'date_modified',
                            'ip_address',
                            ])

        # TODO remove unknown columns, to be checked
        df = df.drop(columns=['Name',
                            'Boss Name'])

        # Rename columns
        df = df.rename(columns={
            'Gebe bitte an, wie zufrieden du bist als Angestellte/r von digitec/Galaxus. Die Skala geht von 1 (schlechtester Wert) bis 10 (bester Wert).Indique ton degré de satisfaction en tant qu’employé(e) digitec/Galaxus. L’échelle va de 1 (la moins bonne note) à 10 (la meilleure note).' : 'Stimmungswert',
            'Was motiviert dich an deinem Job besonders, was trägt besonders zu deiner Zufriedenheit bei?Qu’est-ce qui te motive spécialement dans ton travail, qu’est-ce qui contribue particulièrement à ta satisfaction ?' : 'Motivation 1',
            'Unnamed: 24': 'Motivation 2',
            'Was müsste man verbessern, damit du (noch) zufriedener wärst?Que devrait-on améliorer pour que tu sois (encore) plus satisfait(e)?': 'Verbesserung 1',
            'Unnamed: 26': 'Verbesserung 2',
            'email_address': 'EMAIL',
            'custom_1' : 'VGFIRST',
            'custom_2' : 'VGLAST',
            'custom_3' : 'Abteilung',
            'custom_4' : 'Team',
            'custom_5' : 'Gruppe'
            })

        # Merge all values (unamed columns 14 to 22) into one
        for idx in range (14, 23):
            df['Stimmungswert'] = df['Stimmungswert'].fillna(df[f'Unnamed: {idx}'])

        # Drop the unnamed columns 14 to 22)
        df = df.drop(columns= ['Unnamed: %s' % x for x in range(14,23)])

        # Delte Row 2 some helper text
        df = df.drop(df.index[0])

        df['Stimmungswert'] = df['Stimmungswert'].astype(int)
        df['Vorgesetzter']  = df.apply(lambda row: row['VGLAST'] + ", " + row['VGFIRST'], axis=1)
        df['Mitarbeiter']   = df.apply(lambda row: row['last_name'] + ", " + row['first_name'], axis=1)

        self.df = df

    def _get_id(self, name):

        """
        https://pythex.org
        https://tinyurl.com/yd6p4kbz
        """
        m = re.search('^([A-Z]{2,}[0-9]{0,}|[0-9]{2,})', name)
        if m:
            return m.group(0)
        return name

    def swap_from_id_to_fulltext(self, column, collector):
        """
        Switch back to full names
        :param column: Column name
        :param collector: Collector
        :return: None
        """
        function_name = "get_%s_by_id" % column.lower().replace('-', '')
        function = getattr(collector, function_name)
        self.df[column] =self.df[column].apply(function)

    def swap_from_fulltext_to_id(self, columns):
        """
        Reduce all columns to the unique identifier
        If no ID has been identified, the cell value is kept
        :param columns:
        :return:
        """
        for column in columns:
            self.df[column] = self.df[column].map(self._get_id)

    def workaround_fix_leader_names(self):
        """
        We need to fix some Leader names in 2019.Q2 which are returned with questionmarks from surveymonkey
        """
        # TODO Workaround:
        self.df['Vorgesetzter'] = self.df['Vorgesetzter'].replace("Raki?, Nikola", "Raki\u0107, Nikola")
        self.df['Vorgesetzter'] = self.df['Vorgesetzter'].replace("Mili?i?, Ivana", "Mili\u010di\u0107, Ivana")

    def workaround_add_sub_division_from_collector(self, path, filename):
        """
        Workaround, we have to add sub-division information
        :param path:
        :param filename:
        :return: added sub-division information to dataframe
        """
        # TODO take this information from collector (master.json)

        abt = pd.read_json(os.path.join(path, filename))
        self.df = pd.merge(self.df, abt, on='Mitarbeiter')

    def add_sub_division_company_information(self, dataframejson):
        """
        :param dataframejson: Dataframe "Mitarbeiter", "Sub-Abteilung", "Unternehmen" in JSON format
        :return:
        """
        dfadd = pd.DataFrame.from_dict(dataframejson)
        self.df = pd.merge(self.df, dfadd,  on='Mitarbeiter')

    def get_man_grouped_by_column_as_dict(self, column):
        """
        :param column: column name
        :return: mean value grouped by column as dict
        """
        mean = self.df.groupby(column).mean(numeric_only=True).to_dict()['Stimmungswert']
        return mean

    def add_column_with_the_calulated_mean(self, based_on_columns):
        """
        Add a column with the mean value
        :param based_on_columns: list of columns ot process
        :return:
        """
        for column in based_on_columns:
            self.df[f'{column}-Mean'] = self.df[column]
            mean = self.get_man_grouped_by_column_as_dict(column)
            self.df = self.df.replace({f'{column}-Mean': mean})

    def add_column_with_history_mean(self, based_on_columns, history_dict, date):
        """
        for each base column add the history mean value based on the provided date
        :param based_on_columns:
        :param history_dict:
        :return:
        """
        dateid = date.strftime('%Y.%m.%d')
        for column in based_on_columns:
            #print(f'{dateid}')
            #print("-----------------")
            #print(history_dict[column])
            self.df[f'{dateid}'] = self.df[column]
            #print(self.df[f'{dateid}'])

            self.df[f'{dateid}'] = self.df[f'{dateid}'].map(history_dict[column])
            #print(self.df[f'{dateid}'])

    def add_columns_with_subscriber_statistics(self, based_on_columns, max_subscriber_dict):
        """
        Based on the provided Column (filter) subscriber statistics are added
        in the format of additional columns prefixed with the base column name
        :param based_on_column: list of columns to process
        :param max_subscriber_dict: dict containing the max subscriber inforamtion
        :return:
        """
        for column in based_on_columns:
            self.df[f'{column}-Count'] = self.df[column]
            self.df[f'{column}-Max']   = self.df[column]
            count = self.df.groupby(column).count()['Stimmungswert'].to_dict()
            self.df = self.df.replace({f'{column}-Count': count})
            self.df = self.df.replace({f'{column}-Max': max_subscriber_dict[column]})
            self.df[f'{column}-%'] = self.df.loc[:,f'{column}-Count'].astype(int)\
                                   / self.df.loc[:,f'{column}-Max'].astype(int)

    def drop_duplicates(self):
        """
        Drops the all duplicates
        :return: None
        """
        self.df = self.df.drop_duplicates()

    def get_copy(self, owner = None, collector = None):
        """
        :param owner: owner of the reports
        :return:
        """
        name = self.name
        df = self.df.copy(deep=True)
        if owner:
            name = owner
            leaders = collector.get_list_of_leader_by_top(owner)
            df = df.loc[df['Vorgesetzter'].isin(leaders)]

        return Survey(self.date, df, name)

    def sort(self, columnname):
        """
        Sorts the dataframe
        :param columnname: Name of to columen to be sorted
        :return:
        """
        if "Stimmungswert" in self.df.columns.values.tolist():
            self.df = self.df.sort_values([columnname, 'Stimmungswert'],
                                          ascending=[True, False])
        else:
            self.df =  self.df.sort_values([columnname])

class Collector:

    def __init__(self):
        """
        This class manages all collector information
        """
        self.master = {}

    def read_collector_information_from_jason(self, path, filename):
        """
        Read the collector information from the json file
        :param path: path of the file
        :param filename: filnname
        :return: None
        """
        with open(os.path.join(path, filename)) as json_file:
            self.master = json.load(json_file)

    def _rec_add_to(self, name, res, vglst):
        """
        Create recursive list
        :param name:
        :param res:
        :param vglst:
        :return:
        """
        res.append(name)
        for subname in vglst[name]:
            self._rec_add_to(subname, res, vglst)
        return res

    def get_list_of_leader_by_top(self, top_leader):
        """

        :param top_leader:
        :return:
        """
        ret = []
        self._rec_add_to(top_leader, ret, self.master['tree'])
        return ret

    def check_leader_min_span(self, leader, min_span = 5):
        """

        :param leader:
        :param min_span:
        :return:
        """
        if self.master['span'][leader]['leader'] == 0:
            if self.master['span'][leader]['staff'] < min_span:
                return False
        return True

    def get_vorgesetzter_by_id(self, name):
        return name

    def get_gruppe_by_id(self, idx):
        return self.get_layer_by_id(idx, 'Gruppe')

    def get_abteilung_by_id(self, idx):
        return self.get_layer_by_id(idx, 'Abteilung')

    def get_team_by_id(self, idx):
        return self.get_layer_by_id(idx, 'Team')

    def get_subabteilung_by_id(self, idx):
        return self.get_layer_by_id(idx, 'Sub-Abteilung')

    def get_unternehmen_by_id(self, idx):
        return self.get_layer_by_id(idx, 'Unternehmen')

    def get_layer_by_id(self, idx, layer):
        ret = f'{idx}'
        for name in self.master['id'][layer][idx]:
            ret = ret + " | " + name
        return ret



class History:

    def __init__(self, date):
        """
        Manages history information
        :param date: date of the current report
        """
        self.history = {}
        self.date = date
        self.history['current'] = {}

    def read_history_data_from_json(self, path, filename, date):
        """
        Reads history data
        :param path:
        :param filename:
        :param date: Timestamp (date) of the history file as string (YYYY-mm-dd)
        :return:
        """
        with open(os.path.join(path, filename)) as json_file:
            print(f"> import: {filename}")
            self.history[date] = json.load(json_file)

    def write_own_history_to_json(self, path):
        """
        Exports the history of the current report into an json file
        :param path: Path where the file shall be stored
        :return:
        """
        # Store the history date into an external file for later use
        filename = os.path.join(path, f"history-{self.date.strftime('%Y-%m-%d')}.json")
        with open(filename, "w") as historyfile:
            json.dump(self.history['current'], historyfile, indent=4, sort_keys=True)

    def add_collection(self, name, values):
        """
        :param name: Name of the collector (filter)
        :param values: mean values (dictionary)
        :return:
        """
        # create dictionary for history
        self.history['current'][name] = {}
        self.history['current'][name] = values

    def get_entries_as_sorted_list(self):
        ret = sorted(self.history.keys(), reverse=True)
        # Skip the "current"
        return ret[1:]

    def import_history(self, path):
        """
        Import all history files
        :param path:
        :return:
        """

        files = [filename for filename in os.listdir(path) if filename.startswith("history-")]

        """
        Regex
        https://tinyurl.com/y4hrlyxb
        """
        for f in files:
            m = re.search('history-(20[0-9]{2}-[0-1][0-9]-01).json', f)
            if m:
                date = datetime.strptime(m.group(1), '%Y-%m-%d')
                if date == self.date:
                    print(f'> history import: skip {date}')
                    break
                self.read_history_data_from_json(path, f, date.strftime('%Y.%m.%d'))


    def update_history_fies_from_xlsx(self, path, filename):
        """
        HELPER function only
        :param path:
        :param filename:
        :return:
        """
        wb = load_workbook(os.path.join(path, filename))
        sheet = wb.active
        max_row = sheet.max_row
        max_col = sheet.max_column
        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=max_row, max_col=max_col):
            filter = row[0].value
            key = row[1].value
            date = row[2].value
            mean = row[3].value
            #date = datetime.strptime(row[2].value, '%d.%m.%Y')
            filename = f"history-{date.strftime('%Y-%m-%d')}.json"
            h = {}
            with open(os.path.join(path, filename)) as json_file:

                h = json.load(json_file)
                if "Vorgesetzter" not in h: h["Vorgesetzter"] = {}
                if "Team" not in h: h["Team"] = {}
                if "Sub-Abteilung" not in h: h["Sub-Abteilung"] = {}
                if "Gruppe" not in h: h["Gruppe"] = {}
                if "Unternehmen" not in h: h["Unternehmen"] = {}
                if "Abteilung" not in h: h["Abteilung"] = {}
            h[filter][key] = mean
            #pprint(h)
            #print(f"> write: {filename}")
            with open(os.path.join(path, filename), "w") as json_file:
                json.dump(h, json_file, indent=4)


class Sheet:

    def __init__(self, name):
        """
        """
        self.sheet = name


class Basic(Sheet):

    def __init__(self, survey, filter, writer, collector, history, options):
        """
        """
        self.survey = survey
        self.sheet = filter
        self.writer = writer
        self.options = options
        self.main_col = filter
        self.collector = collector
        self.history = history

        self.check_max_response()

        self.col_idx = [
            self.main_col,
            f'{self.main_col}-Mean',
            f'{self.main_col}-Count',
            f'{self.main_col}-Max',
            f'{self.main_col}-%',
            'Veränderung'
        ]

    def add_column(self, index, columnname):
        """

        :param index:
        :param columnname:
        :return:
        """
        self.col_idx.insert(index, columnname)
        try:
            self.survey.swap_from_id_to_fulltext(columnname, self.collector)
        except:
            pass


    def check_max_response(self):
        """

        :return:
        """
        # Remove all rows where the possible max response is too low
        self.survey.df = self.survey.df.loc[
            self.survey.df[f'{self.main_col}-Max'] > self.options.min_nr_of_resp
            ]



    def set_basic_columns(self):
        """
        Set final index of the columns (column order)
        :return:
        """
        self.survey.df = self.survey.df.reindex(self.col_idx, axis=1)

        # Remove all rows where the possible max response is too low
        #self.survey.df = self.survey.df.loc[
        #    self.survey.df[f'{self.main_col}-Max'] > self.options.min_nr_of_resp
        #]
        #
        #self.add_history()
        #self.finalize()


    def add_history(self):

        for date in self.history.get_entries_as_sorted_list():
            #print(date)
            self.survey.add_column_with_history_mean(
                based_on_columns=[self.sheet],
                history_dict=self.history.history[date],
                date=datetime.strptime(date, '%Y.%m.%d')
            )

        self.survey.df['Veränderung'] = self.survey.df[f'{self.main_col}-Mean'] - \
                                        self.survey.df['2019.01.01']

    def finalize(self):

        #print(f"> finalize {self.sheet}")

        # Fix column naming
        self.survey.df = self.survey.df.rename({
            f'{self.main_col}-%':      'Beteiligung',
            f'{self.main_col}-Mean':   'Mittelwert',
            f'{self.main_col}-Count':  'Anzahl',
            f'{self.main_col}-Max':    'Max',
        }, axis='columns')

        # Sort, remove duplicates, switch to full text names
        self.survey.sort(self.main_col)
        self.survey.drop_duplicates()
        self.survey.swap_from_id_to_fulltext(self.main_col, self.collector)

        # Remove all columns witch are fully empty
        non_null_columns = [col for col in self.survey.df.columns
                            if self.survey.df.loc[:, col].notna().any()]
        self.survey.df = self.survey.df.reindex(non_null_columns, axis=1)


        self.last_row = self.survey.df.shape[0] - 1
        self.last_col = self.survey.df.shape[1] - 1
        self.col_idx  = self.survey.df.columns.values.tolist()


    def write(self):
        """
        Write the dataframe to the sheet
        :return:
        """

        # check if sheet is emmpty
        if not self.col_idx:
            print("> sheet is empty")
            return False

        self.survey.df.to_excel(self.writer, self.sheet)


        # define and set number formats
        workbook  = self.writer.book
        worksheet = self.writer.sheets[self.sheet]

        #https://xlsxwriter.readthedocs.io/worksheet.html#autofilter
        worksheet.autofilter(0, 0, self.last_row + 1, self.last_col + 1)

        text_format = workbook.add_format()
        text_format.set_text_wrap()
        part_format = workbook.add_format({'num_format': '0%'})
        floa_format = workbook.add_format({'num_format': '#,##0.00'})

        if "Veränderung" in self.col_idx:
            worksheet.conditional_format(0,
                                     self.col_idx.index('Veränderung') + 1,
                                     self.last_row + 1,
                                     self.col_idx.index('Veränderung') + 1,
                                     {'type': 'data_bar',
                                      'data_bar_2010': True
                                     })

            worksheet.set_column(
                    self.col_idx.index('Veränderung') + 1,
                    self.col_idx.index('Veränderung') + 1,
                    width = 10,
                    cell_format = floa_format
                    )

        worksheet.set_column(
                    self.col_idx.index(self.main_col) + 1,
                    self.col_idx.index(self.main_col) + 1,
                    width = 40,
                    cell_format = text_format
                    )

        if "Beteiligung" in self.survey.df.columns.values.tolist():
            worksheet.set_column(
                    self.col_idx.index('Beteiligung') + 1,
                    self.col_idx.index('Beteiligung') + 1,
                    width = 10, cell_format = part_format
                    )

        try:
            worksheet.set_column(
                self.col_idx.index(self.history.get_entries_as_sorted_list()[0]),
                self.last_col + 1,
                width = 10,
                cell_format = floa_format
            )
        except Exception as e:
            #print(self.history.get_entries_as_sorted_list())
            #pprint(e)
            #print("<exepion> set float")
            pass

        if "Mittelwert" in self.survey.df.columns.values.tolist():
            worksheet.set_column(
                    self.col_idx.index('Mittelwert') + 1,
                    self.col_idx.index('Mittelwert') + 1,
                    width = 10,
                    cell_format = floa_format
                    )

    def add_vg_abteilung(self):
        """
        TODO A big HACK
        :return:
        """
        d = self.collector.master['ma-to-abt'].copy()
        d['Vorgesetzter'] = d.pop("Mitarbeiter")
        df = pd.DataFrame.from_dict(d)
        self.survey.df = pd.merge(self.survey.df, df, on="Vorgesetzter")

        # FIX ME y (merge)
        # pprint(self.survey.df.columns.values.tolist())

        self.add_column(1, "Sub-Abteilung_y")

class ReportFeedback(Basic):

    def __init__(self, survey, writer, collector, history, options, main_col = "Gruppe", sheet="Feedback Gruppe"):
        """
        """
        self.survey = survey
        self.sheet = sheet
        self.writer = writer
        self.options = options
        self.main_col = main_col
        self.collector = collector
        self.history = history
        self.check_max_response()

        self.col_idx = [
            self.main_col,
            "Stimmungswert",
            "Motivation 1",
            "Motivation 2",
            "Verbesserung 1",
            "Verbesserung 2"
        ]

    def set_formats(self):
        """

        :return:
        """
        # check if sheet is emmpty
        if not self.col_idx:
            print("> sheet is empty")
            return False

        # define and setP number formats
        workbook  = self.writer.book
        worksheet = self.writer.sheets[self.sheet]

        text_format = workbook.add_format()
        text_format.set_text_wrap()

        try:
            worksheet.set_column(
                self.col_idx.index('Motivation 1') + 1,
                self.col_idx.index('Verbesserung 2') + 1,
                width = 40,
                cell_format = text_format
                )
        except:
            pass

        colors = [  "#F8696B",
                    "#FBAA77",
                    "#FFEB84",
                    "#E9E583",
                    "#D3DF82",
                    "#BDD881",
                    "#A6D27F",
                    "#90CB7E",
                    "#7AC57D",
                    "#63BE7B"
                 ]

        for idx in range(0, len(colors)):
            idx_format = workbook.add_format({'bg_color': colors[idx]})
            worksheet.conditional_format(0,
                                     self.col_idx.index('Stimmungswert') + 1,
                                     self.last_row + 1,
                                     self.col_idx.index('Stimmungswert') + 1,
                                     {'type'    : 'cell',
                                      'criteria': "=",
                                      'value'   : idx + 1,
                                      'format'  : idx_format
                                     })

class Process:


    def __init__(self, options):
        """
        """
        self.options = options

        # Set directory where the file is located
        path = os.path.abspath(options.filename)
        self.path = os.path.dirname(path)
        self.filename = os.path.basename(path)

        #print(f'> set working path to: {self.path}')
        #print(f'> set file: {self.filename}')

        self.h = History(datetime.strptime(options.date, '%Y.%m.%d'))

        # FIXME
        self.history   = self.h.history['current']

        self.c = Collector()
        self.c.read_collector_information_from_jason(self.path, 'master.json')

        filters = ['Vorgesetzter', 'Gruppe', 'Team', 'Sub-Abteilung', 'Abteilung', 'Unternehmen']

        self.s = Survey(datetime.strptime(options.date, '%Y.%m.%d'))
        self.s.read_data_from_csv(self.path, self.filename)
        self.s.workaround_fix_leader_names()
        self.s.add_sub_division_company_information(self.c.master['ma-to-abt'])
        self.s.swap_from_fulltext_to_id(filters)
        self.s.add_columns_with_subscriber_statistics(filters, self.c.master['counts'])
        self.s.add_column_with_the_calulated_mean(filters)

    def get_id(self, name):

        """
        https://pythex.org
        https://tinyurl.com/yd6p4kbz
        """
        m = re.search('^([A-Z]{2,}[0-9]{0,}|[0-9]{2,})', name)
        if m:
            return m.group(0)
        return name

    def write_df_sheet(self, df, col_idx, writer, sheet, filter):

        # Write dataframe to the xlsx sheet
        # --------------------------------------------------------------

        # add data frame to sheet
        df.to_excel(writer, sheet)

        last_row = df.shape[0] - 1
        last_col = df.shape[1] - 1

        # define and set number formats
        workbook  = writer.book
        worksheet = writer.sheets[sheet]

        #https://xlsxwriter.readthedocs.io/worksheet.html#autofilter
        worksheet.autofilter(0, 0, last_row + 1, last_col + 1)

        text_format = workbook.add_format()
        text_format.set_text_wrap()
        part_format = workbook.add_format({'num_format': '0%'})
        floa_format = workbook.add_format({'num_format': '#,##0.00'})

        """
        TODO FIXME
        worksheet.set_column(col_idx.index(f'{filter}-%') + 1,
                    col_idx.index(f'{filter}-%') + 1,
                    width = 10, cell_format = part_format)

        worksheet.set_column(col_idx.index(f'{filter}-Mean') + 1,
                            col_idx.index(f'{filter}-Mean') + 1,
                            width = 10,
                            cell_format = floa_format
                            )
        """
        worksheet.set_column(col_idx.index('Motivation 1') + 1,
                            col_idx.index('Verbesserung 2') + 1,
                            width = 40,
                            cell_format = text_format
                            )

        colors = [  "#F8696B",
                    "#FBAA77",
                    "#FFEB84",
                    "#E9E583",
                    "#D3DF82",
                    "#BDD881",
                    "#A6D27F",
                    "#90CB7E",
                    "#7AC57D",
                    "#63BE7B"
                 ]

        for idx in range(0, len(colors)):
            idx_format = workbook.add_format({'bg_color': colors[idx]})
            worksheet.conditional_format(0,
                                     col_idx.index('Stimmungswert') + 1,
                                     last_row + 1,
                                     col_idx.index('Stimmungswert') + 1,
                                     {'type'    : 'cell',
                                      'criteria': "=",
                                      'value'   : idx + 1,
                                      'format'  : idx_format
                                     })





    def individual_report(self, df, vg, filters):

        if not self.c.check_leader_min_span(vg):
            print(f'<<< remove {vg}')
            return False

        s = self.s.get_copy(vg, self.c)

        # open the XLSX writer
        filename = os.path.join(self.path, self.c.master['filenames'][vg])
        writer = pd.ExcelWriter(filename, engine='xlsxwriter')

        division = ReportFeedback(s.get_copy(),writer, self.c, self.h, self.options)
        division.set_basic_columns()
        division.finalize()
        division.write()
        division.set_formats()

        division = ReportFeedback(s.get_copy(),writer, self.c, self.h, self.options, main_col="Vorgesetzter", sheet="Feedback direkt unterstellte MA")
        division.set_basic_columns()
        division.finalize()
        division.write()
        division.set_formats()

        division = Basic(s.get_copy(), "Vorgesetzter", writer, self.c, self.h, self.options)
        division.add_vg_abteilung()
        division.set_basic_columns()
        division.add_history()
        division.finalize()
        division.write()

        division = Basic(s.get_copy(), "Gruppe", writer, self.c, self.h, self.options)
        division.add_column(1, "Abteilung")
        division.set_basic_columns()
        division.add_history()
        division.finalize()
        division.write()

        division = Basic(s.get_copy(), "Team", writer, self.c, self.h, self.options)
        division.add_column(1, "Abteilung")
        division.set_basic_columns()
        division.add_history()
        division.finalize()
        division.write()

        division = Basic(s.get_copy(), "Sub-Abteilung", writer, self.c, self.h, self.options)
        division.set_basic_columns()
        division.add_history()
        division.finalize()
        division.write()

        division = Basic(s.get_copy(), "Abteilung", writer, self.c, self.h, self.options)
        division.set_basic_columns()
        division.add_history()
        division.finalize()
        division.write()

        division = Basic(s.get_copy(), "Unternehmen", writer, self.c, self.h, self.options)
        division.set_basic_columns()
        division.add_history()
        division.finalize()
        division.write()

        # final save
        writer.save()


    def export_history(self):

        filters = ['Vorgesetzter', 'Gruppe', 'Team', 'Sub-Abteilung', 'Abteilung', 'Unternehmen']

        # Make history file
        for name in filters:
            self.h.add_collection(name, self.s.get_man_grouped_by_column_as_dict(name))
        self.h.write_own_history_to_json(self.path)

    def run(self):

        filters = ['Vorgesetzter', 'Gruppe', 'Team', 'Sub-Abteilung', 'Abteilung', 'Unternehmen']

        if options.ceo_only == True:
            self.individual_report(self.s.df, self.c.master["ceo"], filters)
        else:
            for vg in self.c.master['tree']:
                if vg == "NaN":
                    continue
                self.individual_report(self.s.df, vg, filters)



    def tokenzie(self, text):
        #print("---------------")
        #print(text)
        from nltk.tokenize import word_tokenize

        tokenized_word=word_tokenize("why is this")
        #print(tokenized_word)

    def nlp(self):

        df = self.read_data_from_csv()

        df = df.drop(columns=['VGFIRST',
                            'EMAIL',
                            'VGLAST',
                            'Abteilung',
                            'Team',
                            'first_name',
                            'last_name',
                            'Unnamed: 14', 'Unnamed: 15', 'Unnamed: 16',
                            'Unnamed: 17', 'Unnamed: 18', 'Unnamed: 19', 'Unnamed: 20',
                            'Unnamed: 21', 'Unnamed: 22',
                            'Motivation 1', 'Motivation 2'
                            ])

        # Reduce all columnes used as layer to the ID
        #df['Gruppe'] = df['Gruppe'].map(self.get_id)

        df['Verbesserung 1'].map(self.tokenzie)

        #print(df.columns)


        # open the XLSX writer
        writer = pd.ExcelWriter('nlp.xlsx', engine='xlsxwriter')
        df.to_excel(writer, 'x')
        writer.save()

if __name__ == "__main__":

    parser = OptionParser()

    parser.add_option("-v", "--verbose",
        action="store_true", dest="verbose", default="False",
        help="Verbose prints enabled")

    parser.add_option("-c", "--ceo-only",
        action="store_true", dest="ceo_only", default="False",
        help="Only the CEO report")

    parser.add_option("-x", "--exclude",
        type="int", dest="min_nr_of_resp", default="3")

    parser.add_option("-d", "--date", dest="date",
                  help="date of the survey")

    parser.add_option("-f", "--file", dest="filename",
                  help="write report to FILE", metavar="FILE")

    parser.add_option("-m", "--mode", dest="mode",
                  help="vorgesetzter, abteilung, team, gruppe")

    (options, args) = parser.parse_args()

    x = Process(options)
    x.h.import_history(x.path)
    #x.h.get_entries_as_sorted_list()
    x.export_history()
    x.run()

    #x.h.update_history_fies_from_xlsx(x.path, "history.xlsx")
